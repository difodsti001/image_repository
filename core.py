"""
core.py — Motor ColPali + Qdrant + Logica educativa MINEDU
===========================================================
Archivo unificado. Contiene:
  - Carga de modelo ColPali y cliente Qdrant
  - PDF -> imagenes -> embeddings -> insercion en Qdrant
  - Busqueda por texto / imagen
  - Logica educativa: indexacion manual, Excel MINEDU (SSE), busqueda con filtros
  - Verificacion de duplicados contra Qdrant
  - Cancelacion de procesamiento
  - Filtros extraidos directamente de Qdrant
"""

import os
import base64
import logging
import re
import time
import asyncio
import aiohttp
import tempfile
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Optional, AsyncGenerator

import stamina
import torch
from colpali_engine.models import ColPali, ColPaliProcessor
from PIL import Image
from pdf2image import convert_from_bytes, convert_from_path
from qdrant_client import QdrantClient
from qdrant_client.http import models as qm
from dotenv import load_dotenv
import certifi
import ssl

load_dotenv()

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

os.environ["GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"] = certifi.where()
ssl._create_default_https_context = ssl._create_unverified_context

# ---------------------------------------------------------------------------
# Configuracion general
# ---------------------------------------------------------------------------
COLPALI_MODEL_NAME     = os.getenv("COLPALI_MODEL_NAME",     "vidore/colpali-v1.3")
COLPALI_PROCESSOR_NAME = os.getenv("COLPALI_PROCESSOR_NAME", "vidore/colpaligemma-3b-pt-448-base")
PDF_DPI                = int(os.getenv("PDF_DPI",            "300"))
BATCH_SIZE             = int(os.getenv("BATCH_SIZE",         "4"))
QDRANT_URL             = os.getenv("QDRANT_URL",             "http://91.99.108.245:6333")
QDRANT_API_KEY         = os.getenv("QDRANT_API_KEY",         "")
SEARCH_LIMIT           = int(os.getenv("SEARCH_LIMIT",       "5"))
POPPLER_PATH           = os.getenv("POPPLER_PATH",           r"C:\poppler-25.12.0\Library\bin")

# Configuracion educativa
EDU_COLLECTION   = os.getenv("EDU_COLLECTION",    "imagenes_embeddings")
EDU_SEARCH_LIMIT = int(os.getenv("EDU_SEARCH_LIMIT", "5"))
DEVICE = os.getenv("DEVICE", "cpu")  # auto | cpu | cuda

# ---------------------------------------------------------------------------
# Metadata: campos que se guardan en Qdrant por pagina
# ---------------------------------------------------------------------------
# Campos base (siempre presentes)
BASE_PAYLOAD_FIELDS = ["image_base64", "image_name", "source_file", "page_number", "total_pages"]

# Campos de metadata educativa que van en el payload de Qdrant
EDU_META_FIELDS = [
    "categoria", "sub_categoria", "tipo_recurso", "titulo", "modalidad",
    "servicio_educativo", "autor", "derecho_autoridad", "anio_edicion",
    "lengua_idioma", "nivel", "area", "resumen", "competencias",
]

# Campos usados como filtros de busqueda
FILTER_FIELDS = ["tipo_recurso", "nivel", "area", "categoria"]

# Mapeo columna Excel MINEDU -> clave interna
EXCEL_COL_MAP = {
    "CATEGORÍA":          "categoria",
    "SUB CATEGORÍA":      "sub_categoria",
    "TIPO RECURSO":       "tipo_recurso",
    "TÍTULO":             "titulo",
    "MODALIDAD":          "modalidad",
    "SERVICIO EDUCATIVO": "servicio_educativo",
    "AUTOR":              "autor",
    "DERECHO AUTORIDAD":  "derecho_autoridad",
    "AÑO EDICIÓN":        "anio_edicion",
    "LENGUA/IDIOMA":      "lengua_idioma",
    "NIVEL":              "nivel",
    "ÁREA":               "area",
    "RESUMEN":            "resumen",
    "COMPETENCIAS":       "competencias",
    # Campos auxiliares (no van a Qdrant pero se usan para filtrado del Excel)
    "TIPO ENLACE":        "_tipo_enlace",
    "ENLACE":             "_enlace",
    "ESTADO":             "_estado",
}

# ---------------------------------------------------------------------------
# Singletons
# ---------------------------------------------------------------------------
_colpali_model: Optional[ColPali] = None
_colpali_processor: Optional[ColPaliProcessor] = None
_qdrant_client: Optional[QdrantClient] = None
# ---------------------------------------------------------------------------
# Control de cancelacion — usando asyncio.Event para abort real
# ---------------------------------------------------------------------------
_cancel_event: asyncio.Event = asyncio.Event()   # se hace set() para cancelar
_active_task:  asyncio.Task | None = None         # tarea en curso (Excel o manual)

# ---------------------------------------------------------------------------
# Modelo ColPali
# ---------------------------------------------------------------------------

def _detect_device() -> str:
    import sys

    logger.info("Python ejecutándose en: %s", sys.executable)
    logger.info("DEVICE config: %s", DEVICE)

    # 🔥 FORZADO POR CONFIG
    if DEVICE == "cpu":
        logger.info("Forzando CPU por config")
        return "cpu"

    if DEVICE == "cuda":
        if torch.cuda.is_available():
            logger.info("Forzando GPU por config")
            return "cuda"
        logger.warning("CUDA solicitado pero no disponible → CPU")
        return "cpu"

    # AUTO
    logger.info("Modo AUTO")
    logger.info("torch version: %s", torch.__version__)
    logger.info("torch.cuda.is_available(): %s", torch.cuda.is_available())
    logger.info("torch.cuda.device_count(): %s", torch.cuda.device_count())

    if torch.cuda.is_available():
        logger.info("GPU detectada: %s", torch.cuda.get_device_name(0))
        return "cuda"

    logger.info("Sin GPU -> cpu")
    return "cpu"

os.makedirs("./offload", exist_ok=True)
OFFLOAD_DIR = os.getenv("OFFLOAD_DIR", "./offload")

def get_model() -> tuple[ColPali, ColPaliProcessor]:
    global _colpali_model, _colpali_processor

    if _colpali_model and _colpali_processor:
        return _colpali_model, _colpali_processor

    device = _detect_device()
    logger.info("Cargando ColPali '%s' (device=%s)...", COLPALI_MODEL_NAME, device)

    if device == "cpu":
        _colpali_model = ColPali.from_pretrained(
            COLPALI_MODEL_NAME,
            torch_dtype=torch.float32,
            device_map="cpu",
            trust_remote_code=True,
        )

    else:
        _colpali_model = ColPali.from_pretrained(
            COLPALI_MODEL_NAME,
            torch_dtype=torch.float16,
            device_map="auto",
            max_memory={0: "2GiB", "cpu": "14GiB"},
            offload_folder=OFFLOAD_DIR,
            offload_state_dict=True,
            trust_remote_code=True,
        )

    logger.info("Cargando processor '%s'...", COLPALI_PROCESSOR_NAME)
    _colpali_processor = ColPaliProcessor.from_pretrained(COLPALI_PROCESSOR_NAME)

    logger.info("Modelo listo.")
    return _colpali_model, _colpali_processor


def get_qdrant() -> QdrantClient:
    global _qdrant_client
    if _qdrant_client is None:
        logger.info("Conectando a Qdrant en %s...", QDRANT_URL)
        _qdrant_client = QdrantClient(url=QDRANT_URL, api_key=QDRANT_API_KEY or None)
    return _qdrant_client

# ---------------------------------------------------------------------------
# PDF -> imagenes
# ---------------------------------------------------------------------------

def pdf_bytes_to_images(pdf_bytes: bytes, dpi: int = PDF_DPI) -> list[Image.Image]:
    logger.info("Convirtiendo PDF a imagenes (dpi=%d)...", dpi)
    pages = convert_from_bytes(pdf_bytes, dpi=dpi, poppler_path=POPPLER_PATH)
    return [p.convert("RGB") for p in pages]


def image_to_base64(img: Image.Image, max_dim: int = 1600) -> str:
    """
    Serializa imagen PIL a base64 JPEG con calidad reducida.
    Redimensiona si supera max_dim para evitar el limite de 1024KB en Qdrant.
    """
    w, h = img.size
    if max(w, h) > max_dim:
        scale = max_dim / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=75, optimize=True)
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    # Si aun supera 900KB, reducir mas
    if len(b64) > 900_000:
        buf = BytesIO()
        img2 = img.resize((int(img.width * 0.7), int(img.height * 0.7)), Image.LANCZOS)
        img2.save(buf, format="JPEG", quality=65, optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return b64


def base64_to_image(b64: str) -> Image.Image:
    return Image.open(BytesIO(base64.b64decode(b64)))

# ---------------------------------------------------------------------------
# Embeddings
# ---------------------------------------------------------------------------

def embed_images(images: list[Image.Image]) -> list[list[list[float]]]:
    model, processor = get_model()
    all_embeddings = []
    for i in range(0, len(images), BATCH_SIZE):
        batch = images[i: i + BATCH_SIZE]
        logger.info("Embeddings: lote %d-%d de %d...", i+1, i+len(batch), len(images))
        with torch.no_grad():
            inputs = processor.process_images(batch).to(model.device)
            embeddings = model(**inputs)
        for emb in embeddings:
            all_embeddings.append(emb.cpu().float().numpy().tolist())
    return all_embeddings


def embed_query_text(query: str) -> list[list[float]]:
    model, processor = get_model()
    with torch.no_grad():
        inputs = processor.process_queries([query]).to(model.device)
        emb = model(**inputs)
    return emb[0].cpu().float().numpy().tolist()


def embed_query_image(image: Image.Image) -> list[list[float]]:
    model, processor = get_model()
    with torch.no_grad():
        inputs = processor.process_images([image]).to(model.device)
        emb = model(**inputs)
    return emb[0].cpu().float().numpy().tolist()

# ---------------------------------------------------------------------------
# Insercion en Qdrant
# ---------------------------------------------------------------------------

@stamina.retry(on=Exception, attempts=3)
def _upsert_point(client: QdrantClient, collection: str, point: qm.PointStruct) -> None:
    client.upsert(collection_name=collection, points=[point], wait=True)


def index_pdf(
    pdf_bytes: bytes,
    collection_name: str,
    filename: str,
    extra_payload: dict | None = None,
) -> dict:
    """
    Pipeline: PDF -> imagenes -> embeddings -> upsert en Qdrant.
    Chequea cancelacion entre paginas: lanza CancelledError si se pide abort.
    Solo guarda en el payload los campos de EDU_META_FIELDS + BASE_PAYLOAD_FIELDS.
    """
    _check_cancelled()   # antes de empezar

    t0 = time.time()
    images = pdf_bytes_to_images(pdf_bytes)
    n_pages = len(images)

    # Generar embeddings con chequeo de cancelacion por lote
    all_embeddings = []
    model, processor = get_model()
    for i in range(0, len(images), BATCH_SIZE):
        _check_cancelled()
        batch = images[i: i + BATCH_SIZE]
        logger.info("Embeddings: lote %d-%d de %d...", i+1, i+len(batch), len(images))
        with torch.no_grad():
            inputs = processor.process_images(batch).to(model.device)
            embs   = model(**inputs)
        for emb in embs:
            all_embeddings.append(emb.cpu().float().numpy().tolist())

    client  = get_qdrant()
    base_id = int(time.time() * 1000)

    for i, (img, emb) in enumerate(zip(images, all_embeddings)):
        _check_cancelled()   # chequeo entre paginas

        payload: dict = {
            "image_base64": image_to_base64(img),
            "image_name":   f"{Path(filename).stem}_page_{i+1}.jpg",
            "source_file":  filename,
            "page_number":  i + 1,
            "total_pages":  n_pages,
        }
        if extra_payload:
            for field in EDU_META_FIELDS:
                val = extra_payload.get(field)
                if val:
                    payload[field] = val

        try:
            _upsert_point(client, collection_name,
                          qm.PointStruct(id=base_id + i, vector=emb, payload=payload))
        except Exception as exc:
            logger.error("Error insertando pagina %d: %s", i+1, exc)

    elapsed = time.time() - t0
    logger.info("Indexacion: %d paginas en %.2fs", n_pages, elapsed)
    return {"collection": collection_name, "filename": filename,
            "pages_indexed": n_pages, "elapsed_seconds": round(elapsed, 3)}

# ---------------------------------------------------------------------------
# Verificacion de duplicados en Qdrant (sin archivo externo)
# ---------------------------------------------------------------------------

def _normalize_title(title: str) -> str:
    """Normaliza titulo para comparacion (sin acentos, minusculas, sin simbolos)."""
    nfd = unicodedata.normalize("NFD", title.lower())
    clean = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    clean = re.sub(r"[^a-z0-9\s]", "", clean)
    return re.sub(r"\s+", " ", clean).strip()


def check_duplicate_in_qdrant(collection: str, titulo: str) -> tuple[bool, dict]:
    """
    Verifica si ya existe un punto en Qdrant con el mismo titulo (normalizado).
    Hace scroll en la coleccion buscando coincidencias de payload.titulo.
    Retorna (is_duplicate, info_dict).
    """
    if not titulo:
        return False, {}
    client = get_qdrant()
    titulo_norm = _normalize_title(titulo)

    try:
        # Usar filter por campo titulo
        result = client.scroll(
            collection_name=collection,
            scroll_filter=qm.Filter(
                must=[qm.FieldCondition(
                    key="titulo",
                    match=qm.MatchValue(value=titulo)
                )]
            ),
            limit=1,
            with_payload=True,
            with_vectors=False,
        )
        points, _ = result
        if points:
            p = points[0].payload or {}
            return True, {
                "titulo":      p.get("titulo", titulo),
                "source_file": p.get("source_file", ""),
                "page_number": p.get("page_number", 1),
            }
    except Exception:
        pass

    # Fallback: scroll completo con comparacion normalizada (costoso, solo si falla el filtro)
    try:
        offset = None
        while True:
            batch, next_offset = client.scroll(
                collection_name=collection,
                limit=100,
                offset=offset,
                with_payload=["titulo", "source_file", "page_number"],
                with_vectors=False,
            )
            for point in batch:
                p = point.payload or {}
                existing = p.get("titulo", "")
                if existing and _normalize_title(existing) == titulo_norm:
                    return True, {
                        "titulo":      existing,
                        "source_file": p.get("source_file", ""),
                    }
            if next_offset is None:
                break
            offset = next_offset
    except Exception as e:
        logger.warning("check_duplicate fallback error: %s", e)

    return False, {}

# ---------------------------------------------------------------------------
# Filtros desde Qdrant (sin archivo externo)
# ---------------------------------------------------------------------------

def get_filter_options_from_qdrant(collection: str) -> dict:
    """
    Extrae valores unicos de los campos de filtro directamente de Qdrant
    haciendo scroll sobre los payloads.
    Resultado cacheado en memoria durante 5 minutos.
    """
    cache_key = f"_filter_cache_{collection}"
    cached = _filter_cache.get(cache_key)
    if cached and time.time() - cached["ts"] < 300:
        return cached["data"]

    client = get_qdrant()
    tipos, niveles, areas, categorias = set(), set(), set(), set()

    try:
        offset = None
        while True:
            batch, next_offset = client.scroll(
                collection_name=collection,
                limit=250,
                offset=offset,
                with_payload=FILTER_FIELDS,
                with_vectors=False,
            )
            for point in batch:
                p = point.payload or {}
                if p.get("tipo_recurso"): tipos.add(p["tipo_recurso"])
                if p.get("nivel"):        niveles.add(p["nivel"])
                if p.get("area"):         areas.add(p["area"])
                if p.get("categoria"):    categorias.add(p["categoria"])
            if next_offset is None:
                break
            offset = next_offset
    except Exception as e:
        logger.warning("Error obteniendo filtros de Qdrant: %s", e)

    data = {
        "tipos_recurso": sorted(tipos),
        "niveles":       sorted(niveles),
        "areas":         sorted(areas),
        "categorias":    sorted(categorias),
    }
    _filter_cache[cache_key] = {"data": data, "ts": time.time()}
    return data

_filter_cache: dict = {}

# ---------------------------------------------------------------------------
# Busqueda educativa con filtros
# ---------------------------------------------------------------------------

def _build_filter(tipo_recurso=None, nivel=None, area=None, categoria=None):
    conditions = []
    if tipo_recurso:
        conditions.append(qm.FieldCondition(key="tipo_recurso", match=qm.MatchValue(value=tipo_recurso)))
    if nivel:
        conditions.append(qm.FieldCondition(key="nivel", match=qm.MatchValue(value=nivel)))
    if area:
        conditions.append(qm.FieldCondition(key="area", match=qm.MatchValue(value=area)))
    if categoria:
        conditions.append(qm.FieldCondition(key="categoria", match=qm.MatchValue(value=categoria)))
    return qm.Filter(must=conditions) if conditions else None


def _run_search_edu(multivector, limit, tipo_recurso=None, nivel=None, area=None, categoria=None) -> list[dict]:
    client = get_qdrant()
    t0 = time.time()
    result = client.query_points(
        collection_name=EDU_COLLECTION,
        query=multivector,
        limit=limit,
        query_filter=_build_filter(tipo_recurso, nivel, area, categoria),
        timeout=100,
        search_params=qm.SearchParams(
            quantization=qm.QuantizationSearchParams(ignore=False, rescore=True, oversampling=2.0)
        ),
    )
    logger.info("Busqueda en %.4fs — %d resultados", time.time() - t0, len(result.points))

    hits = []
    for point in result.points:
        p = point.payload or {}
        hit = {
            "id":           point.id,
            "score":        point.score,
            "image_base64": p.get("image_base64"),
            "image_name":   p.get("image_name"),
            "source_file":  p.get("source_file"),
            "page_number":  p.get("page_number"),
            "total_pages":  p.get("total_pages"),
        }
        for field in EDU_META_FIELDS:
            hit[field] = p.get(field)
        hits.append(hit)
    return hits


def search_by_text(query, limit=EDU_SEARCH_LIMIT, tipo_recurso=None, nivel=None, area=None, categoria=None):
    logger.info("Busqueda texto: '%s'", query)
    return _run_search_edu(embed_query_text(query), limit, tipo_recurso, nivel, area, categoria)


def search_by_image_query(image: Image.Image, limit=EDU_SEARCH_LIMIT, tipo_recurso=None, nivel=None, area=None, categoria=None):
    logger.info("Busqueda imagen")
    return _run_search_edu(embed_query_image(image), limit, tipo_recurso, nivel, area, categoria)


def search_similar_to_page(image_b64: str, limit=EDU_SEARCH_LIMIT) -> list[dict]:
    """
    Busca paginas similares a una pagina dada en base64.
    Redimensiona la imagen antes de embeddear para evitar el limite de 1024KB.
    """
    logger.info("Busqueda por pagina similar")
    raw = base64.b64decode(image_b64)
    image = Image.open(BytesIO(raw)).convert("RGB")
    # Redimensionar a maximo 800px para que el embedding sea rapido y no exceda limites
    w, h = image.size
    if max(w, h) > 800:
        scale = 800 / max(w, h)
        image = image.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    return _run_search_edu(embed_query_image(image), limit)

# ---------------------------------------------------------------------------
# Indexacion manual
# ---------------------------------------------------------------------------

def index_manual_upload(pdf_bytes: bytes, filename: str, metadata: dict) -> dict:
    """
    Indexa un PDF manualmente. Verifica duplicados en Qdrant antes de procesar.
    Soporta cancelacion via cancel_processing() durante la vectorizacion.
    """
    # Limpiar evento de cancelacion al iniciar
    _cancel_event.clear()

    titulo = metadata.get("titulo") or filename
    is_dup, dup_info = check_duplicate_in_qdrant(EDU_COLLECTION, titulo)
    if is_dup:
        return {
            "error":   True,
            "status":  "skipped",
            "message": (
                f"⚠️ El recurso '{titulo}' ya existe en la base de datos.\n"
                f"Archivo: {dup_info.get('source_file', '')}"
            ),
            "filename": filename,
        }

    try:
        clean_meta = {k: v for k, v in metadata.items() if k in EDU_META_FIELDS and v}
        result = index_pdf(
            pdf_bytes=pdf_bytes,
            collection_name=EDU_COLLECTION,
            filename=filename,
            extra_payload=clean_meta,
        )
        _filter_cache.clear()
        return {**result, **clean_meta}
    except asyncio.CancelledError:
        logger.info("Indexacion manual cancelada: %s", filename)
        return {
            "error":    True,
            "status":   "cancelled",
            "message":  f"⏹ Indexacion cancelada por el usuario ({filename})",
            "filename": filename,
        }

# ---------------------------------------------------------------------------
# Cancelacion del procesamiento Excel
# ---------------------------------------------------------------------------

def cancel_processing():
    """
    Cancela cualquier operacion de indexacion en curso (Excel o manual).
    Hace set() en el evento de cancelacion Y cancela la asyncio.Task activa.
    """
    global _active_task
    _cancel_event.set()
    if _active_task and not _active_task.done():
        _active_task.cancel()
        logger.info("Task de indexacion cancelada.")
    logger.info("Cancelacion solicitada.")


def _check_cancelled():
    """Lanza CancelledError si el usuario pidio cancelacion. Llamar en puntos de chequeo."""
    if _cancel_event.is_set():
        raise asyncio.CancelledError("Operacion cancelada por el usuario")

# ---------------------------------------------------------------------------
# Excel MINEDU: parseo, descarga paralela y vectorizacion con SSE
# ---------------------------------------------------------------------------

def parse_excel_minedu(excel_path: str) -> list[dict]:
    """
    Lee el Excel MINEDU y retorna filas validas:
    - LENGUA/IDIOMA contiene 'castellano'
    - TIPO ENLACE == 'PDF'
    - ESTADO == 'Activo'
    - ENLACE no nulo
    Solo incluye campos de EDU_META_FIELDS en cada entrada.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    raw_headers = [str(c.value).strip() if c.value else ""
                   for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(raw_headers) if h}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def get(col):
            idx = col_idx.get(col)
            if idx is None: return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        lengua   = get("LENGUA/IDIOMA") or ""
        tipo_enl = get("TIPO ENLACE")   or ""
        estado   = get("ESTADO")        or ""
        enlace   = get("ENLACE")

        if "castellano" not in lengua.lower(): continue
        if tipo_enl.upper() != "PDF":          continue
        if estado.lower() != "activo":         continue
        if not enlace:                          continue

        entry = {"_enlace": enlace}
        for excel_col, internal_key in EXCEL_COL_MAP.items():
            if internal_key.startswith("_"):
                continue  # campos auxiliares ya manejados
            val = get(excel_col)
            if val and val != "--":
                entry[internal_key] = val

        rows.append(entry)

    logger.info("Excel: %d filas validas de %d", len(rows), ws.max_row - 1)
    return rows


async def _download_pdf(session: aiohttp.ClientSession, url: str, dest: Path) -> bool:
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=90),
                               ssl=False) as resp:
            if resp.status != 200:
                return False
            dest.write_bytes(await resp.read())
            return True
    except Exception as e:
        logger.warning("Error descargando %s: %s", url, e)
        return False


async def process_excel_minedu(excel_path: str, max_concurrent: int = 4) -> AsyncGenerator[dict, None]:
    """
    Pipeline Excel MINEDU con progreso SSE:
    1. Parsea y filtra filas validas
    2. Verifica duplicados en Qdrant (sin .json)
    3. Descarga PDFs en paralelo
    4. Vectoriza e inserta en Qdrant con metadata correcta
    5. Emite eventos de progreso
    Soporta cancelacion real via cancel_processing() en cualquier momento.
    """
    _cancel_event.clear()   # Reset al iniciar

    rows = parse_excel_minedu(excel_path)
    total = len(rows)

    if total == 0:
        yield {"type": "done", "total": 0, "success": 0, "failed": 0, "skipped": 0,
               "msg": "No se encontraron filas validas (Castellano + PDF + Activo)."}
        return

    yield {"type": "start", "total": total,
           "msg": f"{total} recursos validos encontrados en el Excel"}

    semaphore = asyncio.Semaphore(max_concurrent)
    temp_dir  = Path(tempfile.gettempdir()) / "minedu_downloads"
    temp_dir.mkdir(exist_ok=True)

    success_count = 0
    failed_count  = 0
    skipped_count = 0
    cancelled     = False

    async def process_one(row: dict, pos: int) -> dict:
        # Chequeo inmediato de cancelacion antes de hacer cualquier trabajo
        if _cancel_event.is_set():
            return {"type": "cancelled", "pos": pos, "total": total,
                    "msg": f"[{pos}/{total}] ⏸️ Cancelado"}

        titulo    = row.get("titulo") or f"recurso_{pos}"
        enlace    = row.get("_enlace", "")
        safe_name = re.sub(r"[^a-zA-Z0-9_\-]", "_", titulo[:60]) + ".pdf"

        # Verificar duplicado en Qdrant
        is_dup, _ = check_duplicate_in_qdrant(EDU_COLLECTION, titulo)
        if is_dup:
            return {"type": "skipped", "pos": pos, "total": total,
                    "filename": safe_name, "titulo": titulo,
                    "msg": f"[{pos}/{total}] ⏭️ Omitido (ya existe): {titulo[:50]}"}

        async with semaphore:
            # Segundo chequeo al entrar al semaforo (puede haber esperado)
            if _cancel_event.is_set():
                return {"type": "cancelled", "pos": pos, "total": total,
                        "msg": f"[{pos}/{total}] ⏸️ Cancelado"}

            dest = temp_dir / safe_name
            try:
                async with aiohttp.ClientSession() as session:
                    ok = await _download_pdf(session, enlace, dest)
            except asyncio.CancelledError:
                return {"type": "cancelled", "pos": pos, "total": total,
                        "msg": f"[{pos}/{total}] ⏸️ Descarga cancelada"}

            if not ok:
                return {"type": "error", "pos": pos, "total": total,
                        "filename": safe_name, "titulo": titulo,
                        "msg": f"[{pos}/{total}] ✗ Error descargando: {enlace[:60]}"}

            try:
                pdf_bytes = dest.read_bytes()
                clean_meta = {k: v for k, v in row.items()
                              if k in EDU_META_FIELDS and v}
                # index_pdf ya chequea _cancel_event entre paginas
                result = index_pdf(
                    pdf_bytes=pdf_bytes,
                    collection_name=EDU_COLLECTION,
                    filename=safe_name,
                    extra_payload=clean_meta,
                )
                try: dest.unlink()
                except: pass
                return {"type": "ok", "pos": pos, "total": total,
                        "filename": safe_name, "titulo": titulo,
                        "pages": result["pages_indexed"],
                        "msg": f"[{pos}/{total}] ✓ {titulo[:50]} — {result['pages_indexed']} págs."}
            except asyncio.CancelledError:
                return {"type": "cancelled", "pos": pos, "total": total,
                        "filename": safe_name, "titulo": titulo,
                        "msg": f"[{pos}/{total}] ⏸️ Vectorizacion cancelada: {safe_name}"}
            except Exception as e:
                return {"type": "error", "pos": pos, "total": total,
                        "filename": safe_name, "titulo": titulo,
                        "msg": f"[{pos}/{total}] ✗ Error vectorizando: {e}"}

    tasks = [process_one(row, i+1) for i, row in enumerate(rows)]

    for coro in asyncio.as_completed(tasks):
        event = await coro
        t = event["type"]
        if t == "ok":       success_count += 1
        elif t == "skipped": skipped_count += 1
        elif t == "cancelled":
            cancelled = True
            yield event
            break
        else:
            failed_count += 1
        yield event

    # Invalidar cache de filtros tras indexacion masiva
    _filter_cache.clear()
    _cancel_event.clear()   # reset para siguiente operacion

    yield {
        "type": "done", "total": total,
        "success": success_count, "failed": failed_count,
        "skipped": skipped_count, "cancelled": cancelled,
        "msg": (
            f"Cancelado: {success_count} indexados, {skipped_count} omitidos, {failed_count} fallidos."
            if cancelled else
            f"Completado: {success_count} indexados, {skipped_count} omitidos (ya existian), {failed_count} fallidos."
        ),
    }
