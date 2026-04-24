"""
core_edu.py — Buscador de Recursos Educativos MINEDU
=====================================================
Funciones especificas para el repositorio educativo:
  - Indexacion batch de carpeta local de PDFs
  - Extraccion automatica de metadata desde nombre de archivo
  - Registro de recursos indexados (indice JSON local)
  - Busqueda con filtros (area, nivel, tipo)
  - Feedback de resultados
  - Busqueda por pagina similar (imagen como query)

Depende de core.py para embeddings, Qdrant y modelo ColPali.
"""

import json
import logging
import os
import re
import time
from pathlib import Path
from typing import Optional
from io import BytesIO

from PIL import Image

import core

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuracion especifica del repositorio educativo
# ---------------------------------------------------------------------------
EDU_COLLECTION   = os.getenv("EDU_COLLECTION",   "imagenes_embeddings")
EDU_PDF_FOLDER   = os.getenv("EDU_PDF_FOLDER",   "repositorio")          # carpeta local con los PDFs
EDU_INDEX_FILE   = os.getenv("EDU_INDEX_FILE",   "edu_index.json")       # registro de PDFs indexados
EDU_SEARCH_LIMIT = int(os.getenv("EDU_SEARCH_LIMIT", "5"))

# Vocabularios para extraccion automatica de metadata desde el nombre del archivo
_AREAS = {
    "matematica": "Matemática",
    "math": "Matemática",
    "comunicacion": "Comunicación",
    "comunicación": "Comunicación",
    "ciencia": "Ciencia y Tecnología",
    "cyt": "Ciencia y Tecnología",
    "personal": "Personal Social",
    "social": "Personal Social",
    "religion": "Educación Religiosa",
    "arte": "Arte y Cultura",
    "ingles": "Inglés",
    "english": "Inglés",
    "educacion_fisica": "Educación Física",
    "ef": "Educación Física",
    "dpcc": "DPCC",
    "historia": "Ciencias Sociales",
    "geografia": "Ciencias Sociales",
    "economia": "Ciencias Sociales",
    "fisica": "CTA",
    "quimica": "CTA",
    "biologia": "CTA",
    "cta": "CTA",
    "epт": "EPT",
    "ept": "EPT",
}

_NIVELES = {
    "inicial": "Inicial",
    "primaria": "Primaria",
    "secundaria": "Secundaria",
    "ugel": "Gestión",
    "directivo": "Gestión",
    "docente": "Docente",
}

_TIPOS = {
    "fasciculo": "Fascículo",
    "fascículo": "Fascículo",
    "cuaderno": "Cuaderno de Trabajo",
    "guia": "Guía",
    "guía": "Guía",
    "ficha": "Ficha",
    "modulo": "Módulo",
    "módulo": "Módulo",
    "sesion": "Sesión de Aprendizaje",
    "sesión": "Sesión de Aprendizaje",
    "manual": "Manual",
    "kit": "Kit",
    "presentacion": "Presentación",
    "presentación": "Presentación",
    "informe": "Informe",
    "plan": "Plan",
    "curriculo": "Currículo",
    "curriculo": "Currículo",
    "cneb": "Currículo",
}

_GRADOS = {
    "1ro": "1°", "2do": "2°", "3ro": "3°", "4to": "4°", "5to": "5°", "6to": "6°",
    "1er": "1°", "primer": "1°", "segundo": "2°", "tercero": "3°",
    "cuarto": "4°", "quinto": "5°", "sexto": "6°",
    "1°": "1°", "2°": "2°", "3°": "3°", "4°": "4°", "5°": "5°", "6°": "6°",
    "primero": "1°", "segundo": "2°",
}


# ---------------------------------------------------------------------------
# Extraccion automatica de metadata desde el nombre del archivo
# ---------------------------------------------------------------------------

def extract_metadata_from_filename(filename: str) -> dict:
    """
    Extrae area, nivel, grado y tipo de recurso desde el nombre del archivo.
    Ejemplo: 'Fasciculo_U2S1_Matematica_4to_Primaria.pdf'
             -> {area: 'Matematica', nivel: 'Primaria', grado: '4°', tipo: 'Fasciculo'}
    """
    name = Path(filename).stem.lower()
    # Normalizar separadores
    name_clean = re.sub(r"[_\-\.\s]+", " ", name)
    tokens = name_clean.split()

    area   = next((v for k, v in _AREAS.items()   if k in name_clean), None)
    nivel  = next((v for k, v in _NIVELES.items() if k in name_clean), None)
    tipo   = next((v for k, v in _TIPOS.items()   if k in name_clean), None)
    grado  = next((v for k, v in _GRADOS.items()  if k in tokens), None)

    return {
        "area":   area,
        "nivel":  nivel,
        "grado":  grado,
        "tipo":   tipo,
    }


# ---------------------------------------------------------------------------
# Indice de recursos (JSON local)
# ---------------------------------------------------------------------------

def _load_index() -> dict:
    """Carga el indice JSON de recursos indexados."""
    index_path = Path(EDU_INDEX_FILE)
    if index_path.exists():
        with open(index_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"resources": {}}   # clave: filename, valor: metadata + estado


def _save_index(index: dict) -> None:
    """Guarda el indice JSON."""
    with open(EDU_INDEX_FILE, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)


def get_index() -> dict:
    """Retorna el indice completo de recursos."""
    return _load_index()


def get_resource_list() -> list[dict]:
    """Lista todos los recursos con su estado de indexacion."""
    index = _load_index()
    result = []
    for filename, info in index["resources"].items():
        result.append({
            "filename":    filename,
            "pages":       info.get("pages", 0),
            "indexed":     info.get("indexed", False),
            "indexed_at":  info.get("indexed_at"),
            "area":        info.get("area"),
            "nivel":       info.get("nivel"),
            "grado":       info.get("grado"),
            "tipo":        info.get("tipo"),
        })
    # Ordenar: primero los no indexados, luego por nombre
    result.sort(key=lambda x: (x["indexed"], x["filename"]))
    return result


def get_filter_options() -> dict:
    """
    Retorna los valores unicos disponibles para filtros (area, nivel, tipo, grado)
    basados en los recursos ya indexados.
    """
    index = _load_index()
    areas, niveles, tipos, grados = set(), set(), set(), set()
    for info in index["resources"].values():
        if info.get("indexed"):
            if info.get("area"):   areas.add(info["area"])
            if info.get("nivel"):  niveles.add(info["nivel"])
            if info.get("tipo"):   tipos.add(info["tipo"])
            if info.get("grado"):  grados.add(info["grado"])
    return {
        "areas":   sorted(areas),
        "niveles": sorted(niveles),
        "tipos":   sorted(tipos),
        "grados":  sorted(grados),
    }


# ---------------------------------------------------------------------------
# Carga de metadata desde Excel
# ---------------------------------------------------------------------------

def load_metadata_from_excel(excel_path: str) -> dict:
    """
    Carga metadata de recursos educativos desde un archivo Excel.
    El Excel debe tener una columna 'filename' (nombre del PDF) como campo comun,
    y columnas opcionales: area, nivel, grado, tipo (y cualquier otra).

    Retorna un dict: { filename -> {area, nivel, grado, tipo, ...} }

    Ejemplo de Excel:
      filename                          | area        | nivel    | grado | tipo
      Fasciculo_Mat_4to_Primaria.pdf    | Matematica  | Primaria | 4to   | Fasciculo
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if "filename" not in headers:
        raise ValueError("El Excel debe tener una columna llamada 'filename' con el nombre del PDF.")

    col_idx = {h: i for i, h in enumerate(headers)}
    meta_map = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        fname = row[col_idx["filename"]]
        if not fname:
            continue
        fname = str(fname).strip()

        entry = {}
        for field in ("area", "nivel", "grado", "tipo"):
            if field in col_idx:
                val = row[col_idx[field]]
                entry[field] = str(val).strip() if val else None
            else:
                entry[field] = None

        # Campos extra (cualquier columna adicional se guarda como metadata)
        for h in headers:
            if h not in ("filename", "area", "nivel", "grado", "tipo") and h:
                val = row[col_idx[h]]
                entry[h] = str(val).strip() if val else None

        meta_map[fname] = entry

    logger.info("Metadata cargada desde Excel: %d entradas desde '%s'", len(meta_map), excel_path)
    return meta_map


def apply_excel_metadata(excel_path: str) -> dict:
    """
    Aplica la metadata del Excel al indice de recursos.
    Para cada PDF en el indice que tenga entrada en el Excel,
    sobreescribe area, nivel, grado, tipo (y campos extra).
    Los PDFs sin entrada en el Excel conservan su metadata extraida del nombre.

    Retorna resumen: {updated, not_found_in_index, total_excel_entries}
    """
    meta_map = load_metadata_from_excel(excel_path)
    index = _load_index()

    updated = 0
    not_found = []

    for fname, meta in meta_map.items():
        if fname in index["resources"]:
            index["resources"][fname].update({k: v for k, v in meta.items() if v is not None})
            updated += 1
            logger.info("Metadata actualizada desde Excel: %s", fname)
        else:
            not_found.append(fname)
            logger.warning("PDF del Excel no esta en el indice: %s", fname)

    _save_index(index)
    return {
        "updated":              updated,
        "not_found_in_index":   not_found,
        "total_excel_entries":  len(meta_map),
    }


# ---------------------------------------------------------------------------
# Escaneo y registro de PDFs en la carpeta local
# ---------------------------------------------------------------------------

def scan_repository() -> dict:
    """
    Escanea la carpeta EDU_PDF_FOLDER:
    - Registra PDFs nuevos encontrados en el indice.
    - Elimina del indice los PDFs que ya no existen en disco
      (solo si aun no fueron indexados; los indexados se conservan porque
      sus vectores siguen en Qdrant).
    """
    folder = Path(EDU_PDF_FOLDER)
    if not folder.exists():
        folder.mkdir(parents=True, exist_ok=True)
        logger.info("Carpeta del repositorio creada: %s", folder)

    pdf_files = list(folder.glob("*.pdf")) + list(folder.glob("**/*.pdf"))
    present_names = {p.name for p in pdf_files}
    index = _load_index()

    # Eliminar del indice los PDFs pendientes que ya no existen en disco
    removed_count = 0
    to_remove = [
        fname for fname, info in index["resources"].items()
        if not info.get("indexed") and fname not in present_names
    ]
    for fname in to_remove:
        del index["resources"][fname]
        removed_count += 1
        logger.info("Recurso eliminado del indice (ya no existe): %s", fname)

    # Registrar PDFs nuevos
    new_count = 0
    for pdf_path in pdf_files:
        fname = pdf_path.name
        if fname not in index["resources"]:
            meta = extract_metadata_from_filename(fname)
            index["resources"][fname] = {
                "path":       str(pdf_path.resolve()),
                "indexed":    False,
                "indexed_at": None,
                "pages":      0,
                **meta,
            }
            new_count += 1
            logger.info("Nuevo recurso registrado: %s", fname)
        else:
            # Actualizar path por si cambio de ubicacion
            index["resources"][fname]["path"] = str(pdf_path.resolve())

    _save_index(index)
    total   = len(index["resources"])
    indexed = sum(1 for v in index["resources"].values() if v.get("indexed"))
    return {
        "total":    total,
        "indexed":  indexed,
        "pending":  total - indexed,
        "new_found": new_count,
        "removed":  removed_count,
    }


# ---------------------------------------------------------------------------
# Indexacion batch
# ---------------------------------------------------------------------------

def index_pending(max_files: int = 0) -> dict:
    """
    Vectoriza e inserta en Qdrant todos los PDFs pendientes del repositorio.

    Args:
        max_files: Limite de archivos a procesar en esta llamada (0 = todos).

    Returns:
        Resumen con archivos procesados, paginas insertadas y errores.
    """
    index = _load_index()
    pending = [
        (fname, info) for fname, info in index["resources"].items()
        if not info.get("indexed") and Path(info.get("path", "")).exists()
    ]

    if max_files > 0:
        pending = pending[:max_files]

    if not pending:
        logger.info("No hay PDFs pendientes de indexar.")
        return {"processed": 0, "pages_total": 0, "errors": []}

    processed, pages_total = 0, 0
    errors = []

    for fname, info in pending:
        pdf_path = Path(info["path"])
        logger.info("Indexando: %s", fname)
        try:
            pdf_bytes = pdf_path.read_bytes()
            result = core.index_pdf(
                pdf_bytes=pdf_bytes,
                collection_name=EDU_COLLECTION,
                filename=fname,
                extra_payload={
                    "area":  info.get("area"),
                    "nivel": info.get("nivel"),
                    "grado": info.get("grado"),
                    "tipo":  info.get("tipo"),
                },
            )
            # Actualizar indice
            index["resources"][fname]["indexed"]    = True
            index["resources"][fname]["indexed_at"] = time.strftime("%Y-%m-%dT%H:%M:%S")
            index["resources"][fname]["pages"]      = result["pages_indexed"]
            _save_index(index)

            processed   += 1
            pages_total += result["pages_indexed"]
            logger.info("  OK: %d paginas", result["pages_indexed"])

        except Exception as exc:
            logger.error("  Error indexando %s: %s", fname, exc)
            errors.append({"file": fname, "error": str(exc)})

    return {"processed": processed, "pages_total": pages_total, "errors": errors}


def index_single_upload(pdf_bytes: bytes, filename: str) -> dict:
    """
    Indexa un PDF subido manualmente (no esta en la carpeta local).
    Lo agrega al indice y lo vectoriza de inmediato.
    """
    meta = extract_metadata_from_filename(filename)
    index = _load_index()

    # Registrar en indice (sin path local, fue subido via API)
    index["resources"][filename] = {
        "path":       None,
        "indexed":    False,
        "indexed_at": None,
        "pages":      0,
        **meta,
    }

    result = core.index_pdf(
        pdf_bytes=pdf_bytes,
        collection_name=EDU_COLLECTION,
        filename=filename,
        extra_payload=meta,
    )

    index["resources"][filename]["indexed"]    = True
    index["resources"][filename]["indexed_at"] = time.strftime("%Y-%m-%dT%H:%M:%S")
    index["resources"][filename]["pages"]      = result["pages_indexed"]
    _save_index(index)

    return {**result, **meta}


# ---------------------------------------------------------------------------
# Busqueda con filtros
# ---------------------------------------------------------------------------

def _build_filter(area: str = None, nivel: str = None,
                  tipo: str = None,  grado: str = None):
    """Construye el filtro de Qdrant a partir de los parametros opcionales."""
    from qdrant_client.http import models as qmodels
    conditions = []
    if area:
        conditions.append(qmodels.FieldCondition(
            key="area", match=qmodels.MatchValue(value=area)))
    if nivel:
        conditions.append(qmodels.FieldCondition(
            key="nivel", match=qmodels.MatchValue(value=nivel)))
    if tipo:
        conditions.append(qmodels.FieldCondition(
            key="tipo", match=qmodels.MatchValue(value=tipo)))
    if grado:
        conditions.append(qmodels.FieldCondition(
            key="grado", match=qmodels.MatchValue(value=grado)))
    if not conditions:
        return None
    return qmodels.Filter(must=conditions)


def search_resources_by_text(
    query: str,
    limit: int = EDU_SEARCH_LIMIT,
    area:  str = None,
    nivel: str = None,
    tipo:  str = None,
    grado: str = None,
) -> list[dict]:
    """Busca recursos por texto con filtros opcionales."""
    logger.info("Busqueda educativa por texto: '%s'", query)
    multivector = core.embed_query_text(query)
    return _search_edu(multivector, limit, area, nivel, tipo, grado)


def search_resources_by_image(
    image: Image.Image,
    limit: int = EDU_SEARCH_LIMIT,
    area:  str = None,
    nivel: str = None,
    tipo:  str = None,
    grado: str = None,
) -> list[dict]:
    """Busca recursos usando una imagen como query."""
    logger.info("Busqueda educativa por imagen")
    multivector = core.embed_query_image(image)
    return _search_edu(multivector, limit, area, nivel, tipo, grado)


def search_similar_to_page(image_b64: str, limit: int = EDU_SEARCH_LIMIT) -> list[dict]:
    """
    Busca paginas visualmente similares a una pagina ya existente en el repositorio.
    Recibe el base64 de la imagen y lo usa como query.
    """
    logger.info("Busqueda por pagina similar")
    image = Image.open(BytesIO(__import__("base64").b64decode(image_b64))).convert("RGB")
    multivector = core.embed_query_image(image)
    return _search_edu(multivector, limit)


def _search_edu(
    multivector: list[list[float]],
    limit: int,
    area:  str = None,
    nivel: str = None,
    tipo:  str = None,
    grado: str = None,
) -> list[dict]:
    """Busqueda interna con filtros opcionales sobre la coleccion educativa."""
    from qdrant_client.http import models as qmodels
    client = core.get_qdrant()
    t0 = time.time()

    query_filter = _build_filter(area, nivel, tipo, grado)

    result = client.query_points(
        collection_name=EDU_COLLECTION,
        query=multivector,
        limit=limit,
        query_filter=query_filter,
        timeout=100,
        search_params=qmodels.SearchParams(
            quantization=qmodels.QuantizationSearchParams(
                ignore=False, rescore=True, oversampling=2.0,
            )
        ),
    )

    elapsed = time.time() - t0
    logger.info("Busqueda educativa completada en %.4fs — %d resultados", elapsed, len(result.points))

    hits = []
    for point in result.points:
        p = point.payload or {}
        hits.append({
            "id":           point.id,
            "score":        point.score,
            "image_base64": p.get("image_base64"),
            "image_name":   p.get("image_name"),
            "source_file":  p.get("source_file"),
            "page_number":  p.get("page_number"),
            "total_pages":  p.get("total_pages"),
            "area":         p.get("area"),
            "nivel":        p.get("nivel"),
            "grado":        p.get("grado"),
            "tipo":         p.get("tipo"),
        })
    return hits



